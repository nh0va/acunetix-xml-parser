# Acunetix XML results parser by Ramiro Molina modified by Nicolas Hovassapian
# This script requires openpyxl (https://pypi.python.org/pypi/openpyxl)
# and python 3.x

import os
import sys
import glob
from openpyxl import workbook
from openpyxl import worksheet
from openpyxl import writer
import xml.etree.ElementTree as ET
from html.parser import HTMLParser

class MLStripper(HTMLParser):
    def __init__(self):
        super().__init__()
        self.reset()
        self.fed = []
    def handle_data(self, d):
        self.fed.append(d)
    def get_data(self):
        return ''.join(self.fed)

def file_len(fname):
    i = -1
    with open(fname) as f:
        for i, l in enumerate(f):
            pass
    return i + 1
        
def strip_tags(html):
    s = MLStripper()
    s.feed(html)
    return s.get_data()

def doWork(xml_list,ouputfile):
    files_parsed = 0
    files_skipped = 0
    
    wb = workbook.Workbook()
    ws = wb.create_sheet(title='Vulnerabilities')
    columns = ['UniqueID','Timestamp','StartURL','WebServer','Issue Type','Name','Risk Rating','Description','DetailedInformation','Details','Affects','Impact','Recommendation','References','Request','Response','Is False Positive', 'CVSS Vector', 'CVSS Score','CVSS3 Vector', 'CVSS3 Score']
    for i in range(len(columns)):
        ws.cell(row=1, column=i+1, value=columns[i])
    wb.save(ouputfile)  

    j = 2
    for xml_file in xml_list:
        if file_len(xml_file) != 27: ##For handling empty files (27 lines long)
            print ('Parsing result file: ' + xml_file)
            tree = ET.parse(xml_file)
            root = tree.getroot()    
            
            startURL = root.find('./Scan/StartURL')
            startTime = root.find('./Scan/StartTime')
            webServer =  root.find('./Scan/WebServer')
            
            itemcount = 0
            for reportItem in root.iter('ReportItem'):
                ws.cell(row=j, column=1, value=reportItem.get('id'))
                ws.cell(row=j, column=2, value=startTime.text)
                ws.cell(row=j, column=3, value=startURL.text)
                ws.cell(row=j, column=4, value=webServer.text)
                ws.cell(row=j, column=5, value=reportItem.find('Type').text)
                ws.cell(row=j, column=6, value=reportItem.find('Name').text)
                ws.cell(row=j, column=7, value=reportItem.find('Severity').text)
                ws.cell(row=j, column=8, value=strip_tags(reportItem.find('Description').text))
                detailedInformation = reportItem.find('DetailedInformation')
                if detailedInformation is not None and detailedInformation.text is not None:
                    ws.cell(row=j, column=9, value=strip_tags(detailedInformation.text))
                details = reportItem.find('Details')
                if details is not None and details.text is not None:
                    ws.cell(row=j, column=10, value=strip_tags(details.text))
                ws.cell(row=j, column=11, value=reportItem.find('Affects').text)
                impact = reportItem.find('Impact').text
                if impact is not None:
                    ws.cell(row=j, column=12, value=impact)
                ws.cell(row=j, column=13, value=strip_tags(reportItem.find('Recommendation').text))
                
                #Include references items
                reference = ''
                for referenceItem in reportItem.iter('Reference'):
                    reference = reference + referenceItem.find('Database').text + ': ' + referenceItem.find('URL').text + '\n'
                ws.cell(row=j, column=14, value=reference)
                
                request = reportItem.find('TechnicalDetails/Request').text
                if request is not None:
                    ws.cell(row=j, column=15, value=request)
                response = reportItem.find('TechnicalDetails/Response')
                if response is not None and response.text is not None:
                    ws.cell(row=j, column=16, value=response.text)
                    
                CVSS_vector = reportItem.find('CVSS/Descriptor')
                if CVSS_vector is not None and CVSS_vector.text is not None:
                    ws.cell(row=j, column=18, value=CVSS_vector.text)
                    
                CVSS_score = reportItem.find('CVSS/Score')
                if CVSS_score is not None and CVSS_score.text is not None:
                    ws.cell(row=j, column=19, value=CVSS_score.text)
                
                CVSS3_vector = reportItem.find('CVSS3/Descriptor')
                if CVSS3_vector is not None and CVSS3_vector.text is not None:
                    ws.cell(row=j, column=20, value=CVSS3_vector.text)
                    
                CVSS3_score = reportItem.find('CVSS3/Score')
                if CVSS3_score is not None and CVSS3_score.text is not None:
                    ws.cell(row=j, column=21, value=CVSS3_score.text)
                    
                j = j + 1
                itemcount = itemcount + 1
            print ('Number of elements found in file: ' + str(itemcount))
            files_parsed+=1
        else:
            print ('Skipping empty result file: ' + xml_file)
            files_skipped+=1
    wb.save(ouputfile)
    print ('\n\nDone! Total of files: {0}, total of parsed files: {1}, total of skipped files: {2}'.format(len(xml_list),files_parsed,files_skipped))

def banner():
    print ('Acunetix XML report file parser. Support for Acunetix v11. Author: Ramiro Molina - Modified by Nicolas Hovassapian\n')

def usage(scriptName):
    print ('Proper usage is: python ' + scriptName + ' <Directory with XML Files> <XLSX output file>')
    
if __name__== '__main__':
    try:
        banner()
        if len(sys.argv) == 3:
            if os.path.isdir(sys.argv[1]):
                if sys.argv[1].endswith('\\'):
                    xml_list = glob.glob(sys.argv[1] + '*.xml')
                else:
                    xml_list = glob.glob(sys.argv[1] + '\*.xml')
                if len(xml_list) > 0:
                    print ('Saving results to: ' + sys.argv[2] + '\n')
                    doWork(xml_list,sys.argv[2])
                else:
                    print ('\nError: No XML files were found in the supplied directory: "{0}"\n'.format(sys.argv[1]))
                    sys.exit(-1)
            else:
                print ('Error: "{0}" Is not a valid directory'.format(sys.argv[1]))
                sys.exit(-1)
        else:
            usage(sys.argv[0])
    except KeyboardInterrupt:
        print ('\nBye\n')
        sys.exit(0)